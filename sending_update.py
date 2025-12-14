"""
HCN Email Management System (Automated)
========================================
Flow:
1. Send initial HCN request emails
2. Check inbox & analyze with OpenAI (Received/Critical/Non Critical)
3. Auto-send reminder if no HCN received after 2 hours

Categories:
- Received: HCN found → Update SupplierHCN column
- Critical: Cannot provide HCN (no room, rate issue)
- Non Critical: Other response, no HCN yet
"""

import pandas as pd
import smtplib
import imaplib
import email
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.header import decode_header
from datetime import datetime, timedelta
from openai import OpenAI
import time
import os
import json
import re
from openpyxl import load_workbook

# ========================= CONFIGURATION =========================
# Import all configuration from config.py (which loads from .env file)
from config import (
    GMAIL_ADDRESS,
    GMAIL_APP_PASSWORD,
    OPENAI_API_KEY,
    EXCEL_FILE_PATH,
    SHEET_NAME,
    REMINDER_AFTER_HOURS,
    DAYS_TO_CHECK,
    DELAY_BETWEEN_EMAILS,
    COMPANY_NAME,
    SENDER_NAME
)
# =================================================================


class HCNEmailManager:
    def __init__(self):
        self.gmail_address = GMAIL_ADDRESS
        self.gmail_password = GMAIL_APP_PASSWORD
        self.openai_client = OpenAI(api_key=OPENAI_API_KEY)
        self.excel_path = EXCEL_FILE_PATH
        self.sheet_name = SHEET_NAME
    
    # ==================== EXCEL FUNCTIONS ====================
    
    def read_excel(self):
        """Read bookings from Excel"""
        df = pd.read_excel(self.excel_path, sheet_name=self.sheet_name, header=1)
        df = self.ensure_columns(df)
        return df
    
    def ensure_columns(self, df):
        """Ensure all tracking columns exist"""
        new_columns = {
            'EmailSent': None,
            'EmailSentTime': None,
            'ReminderSent': None,
            'ReminderTime': None,
            'Issue': None
        }
        for col, default in new_columns.items():
            if col not in df.columns:
                df[col] = default
        return df
    
    def save_excel(self, df):
        """Save updated data to Excel"""
        wb = load_workbook(self.excel_path)
        ws = wb[self.sheet_name]
        
        # Get column mapping from header row (row 2)
        columns = {}
        for col_idx, cell in enumerate(ws[2], 1):
            if cell.value:
                columns[cell.value] = col_idx
        
        # Add new columns if not exist
        new_cols = ['EmailSent', 'EmailSentTime', 'ReminderSent', 'ReminderTime', 'Issue']
        max_col = max(columns.values()) if columns else 1
        
        for col_name in new_cols:
            if col_name not in columns:
                max_col += 1
                columns[col_name] = max_col
                ws.cell(row=2, column=max_col, value=col_name)
        
        # Update data rows (data starts from row 3)
        for idx, row in df.iterrows():
            excel_row = idx + 3
            
            for col_name in ['EmailSent', 'EmailSentTime', 'ReminderSent', 'ReminderTime', 'Issue', 'SupplierHCN']:
                if col_name in columns and col_name in df.columns:
                    value = row.get(col_name)
                    if pd.notna(value):
                        ws.cell(row=excel_row, column=columns[col_name], value=value)
        
        wb.save(self.excel_path)
        print(f"   ✅ Excel saved")
    
    # ==================== EMAIL FUNCTIONS ====================
    
    def get_recipient_email(self, row):
        """Get recipient email from row"""
        agent_email = row.get('Agent Email', '')
        if pd.notna(agent_email) and str(agent_email).strip() and '@' in str(agent_email):
            return str(agent_email).strip()
        return None
    
    def create_email_content(self, row, is_reminder=False):
        """Create email subject and body"""
        guest_name = row.get('GuestName', 'Guest')
        hotel_name = row.get('HotelName', 'Hotel')
        from_date = row.get('FromDate', '')
        to_date = row.get('ToDate', '')
        room_type = row.get('RoomType', '')
        no_of_rooms = row.get('NoOFRooms', 1)
        no_of_pax = row.get('NoOfPax', '')
        file_no = row.get('FileNo', '')
        supplier_ref = row.get('SupplierRef', '')
        supplier_name = row.get('SupplierName', '')
        city_name = row.get('CityName', '')
        country_name = row.get('CountryName', '')
        
        # Format dates
        if pd.notna(from_date) and isinstance(from_date, datetime):
            from_date = from_date.strftime('%d-%b-%Y')
        if pd.notna(to_date) and isinstance(to_date, datetime):
            to_date = to_date.strftime('%d-%b-%Y')
        
        prefix = "REMINDER: " if is_reminder else ""
        urgency = "\n⚠️ REMINDER: We have not received the HCN for this booking yet.\n" if is_reminder else ""
        
        subject = f"{prefix}HCN Request - {guest_name} | {hotel_name} | Ref: {file_no}"
        
        body = f"""Dear Team,

Greetings!
{urgency}
We kindly request the Hotel Confirmation Number (HCN) for the following booking:

BOOKING DETAILS:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Guest Name       : {guest_name}
Hotel Name       : {hotel_name}
Location         : {city_name}, {country_name}
Check-in Date    : {from_date}
Check-out Date   : {to_date}
Room Type        : {room_type}
No. of Rooms     : {no_of_rooms}
No. of Guests    : {no_of_pax}
Supplier         : {supplier_name}
Our Reference    : {file_no}
Supplier Ref     : {supplier_ref if pd.notna(supplier_ref) and str(supplier_ref).strip() else 'N/A'}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Please reply with the Hotel Confirmation Number (HCN).

Best Regards,
{SENDER_NAME}
{COMPANY_NAME}

Reference: {file_no}
"""
        return subject, body
    
    def send_email(self, recipient, subject, body):
        """Send email via Gmail SMTP"""
        try:
            msg = MIMEMultipart()
            msg['From'] = self.gmail_address
            msg['To'] = recipient
            msg['Subject'] = subject
            msg.attach(MIMEText(body, 'plain'))
            
            with smtplib.SMTP('smtp.gmail.com', 587) as server:
                server.starttls()
                server.login(self.gmail_address, self.gmail_password)
                server.send_message(msg)
            return True, "Sent"
        except Exception as e:
            return False, str(e)
    
    # ==================== GMAIL IMAP ====================
    
    def connect_gmail_imap(self):
        """Connect to Gmail IMAP"""
        try:
            mail = imaplib.IMAP4_SSL('imap.gmail.com')
            mail.login(self.gmail_address, self.gmail_password)
            mail.select('inbox')
            return mail
        except Exception as e:
            print(f"❌ Gmail error: {str(e)}")
            return None
    
    def decode_header_value(self, header):
        """Decode email header"""
        if not header:
            return ""
        decoded = decode_header(header)
        result = ""
        for part, enc in decoded:
            if isinstance(part, bytes):
                result += part.decode(enc or 'utf-8', errors='ignore')
            else:
                result += part
        return result
    
    def get_email_body(self, msg):
        """Extract email body"""
        body = ""
        if msg.is_multipart():
            for part in msg.walk():
                if part.get_content_type() == "text/plain":
                    try:
                        payload = part.get_payload(decode=True)
                        if payload:
                            body += payload.decode('utf-8', errors='ignore')
                    except:
                        pass
        else:
            try:
                payload = msg.get_payload(decode=True)
                if payload:
                    body = payload.decode('utf-8', errors='ignore')
            except:
                body = str(msg.get_payload())
        return body
    
    # ==================== OPENAI ANALYSIS ====================
    
    def analyze_with_openai(self, subject, body, booking_info):
        """
        Use OpenAI API to analyze email content and categorize:
        - Received: HCN number is provided
        - Critical: Cannot provide HCN (no room, rate issue, sold out, etc.)
        - Non Critical: Everything else
        """
        try:
            our_reference = booking_info.get('file_no', '')
            supplier_ref = booking_info.get('supplier_ref', '')
            
            prompt = f"""You are analyzing a hotel booking reply email to extract the HOTEL CONFIRMATION NUMBER (HCN).

=== CRITICAL: DO NOT CONFUSE THESE ===
OUR REFERENCE (NOT HCN): {our_reference}
SUPPLIER REFERENCE (NOT HCN): {supplier_ref}

These are OUR booking references that WE sent to them. DO NOT extract these as HCN.
The HCN is a NEW number that the HOTEL/SUPPLIER provides in their REPLY.

=== BOOKING DETAILS ===
Guest: {booking_info.get('guest_name')}
Hotel: {booking_info.get('hotel_name')}

=== EMAIL TO ANALYZE ===
SUBJECT: {subject}

BODY:
{body[:3000]}

=== YOUR TASK ===
1. Check if this email contains a HOTEL CONFIRMATION NUMBER (HCN)
2. The HCN must be a NEW number from the hotel/supplier, NOT our reference numbers
3. Categorize the email

=== WHAT IS AN HCN? ===
- A confirmation number PROVIDED BY THE HOTEL in their reply
- Usually labeled as: "Confirmation Number", "Conf#", "HCN", "Hotel Confirmation", "Booking ID", "Reservation ID"
- It's a number/code the HOTEL gives to confirm they received the booking

=== WHAT IS NOT AN HCN? ===
- Our reference number: {our_reference} - NEVER extract this as HCN
- Supplier reference: {supplier_ref} - NEVER extract this as HCN
- Any reference WE mentioned in our original request
- File numbers starting with patterns like: OSTR, DIDA, similar prefixes from our system
- Phone numbers, dates, room numbers, rate amounts

=== CATEGORY RULES ===
1. "Received" = Email contains a VALID NEW HCN from the hotel (not our reference)
2. "Critical" = Email says they CANNOT confirm:
   - No room available / Sold out / Fully booked
   - Rate issue / Price not available / Rate mismatch
   - Cannot confirm / Booking rejected / Declined
   - Cancelled / Not possible
   NOTE: If Critical, HCN should be null
3. "Non Critical" = Everything else:
   - Will check and revert
   - Processing your request
   - Acknowledged / Received
   - No clear HCN and no critical issue

Respond with ONLY this JSON:
{{
    "hcn_number": "NEW confirmation number from hotel, or null",
    "category": "Received" or "Critical" or "Non Critical",
    "reason": "brief explanation"
}}"""

            response = self.openai_client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": "You analyze hotel emails and extract HCN numbers. Respond with JSON only."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.1,
                max_tokens=300
            )
            
            result_text = response.choices[0].message.content.strip()
            
            # Clean JSON
            if '```' in result_text:
                result_text = re.sub(r'^```json?\s*', '', result_text)
                result_text = re.sub(r'\s*```$', '', result_text)
            
            result = json.loads(result_text)
            
            # VALIDATION: Make sure HCN is not our reference number
            hcn = result.get('hcn_number')
            if hcn:
                hcn_str = str(hcn).strip()
                hcn_lower = hcn_str.lower()
                our_ref_lower = str(our_reference).lower().strip()
                supplier_ref_lower = str(supplier_ref).lower().strip()
                
                is_invalid = False
                
                # Check if HCN matches our references
                if our_ref_lower and (hcn_lower == our_ref_lower or hcn_lower in our_ref_lower or our_ref_lower in hcn_lower):
                    is_invalid = True
                    result['reason'] = f"Rejected: '{hcn_str}' matches our FileNo. " + result.get('reason', '')
                
                elif supplier_ref_lower and (hcn_lower == supplier_ref_lower or hcn_lower in supplier_ref_lower or supplier_ref_lower in hcn_lower):
                    is_invalid = True
                    result['reason'] = f"Rejected: '{hcn_str}' matches SupplierRef. " + result.get('reason', '')
                
                # Check for common internal reference patterns
                internal_patterns = ['OSTR', 'DIDA', 'OTLMA', 'DIDAMA', 'FILE-', 'REF-', 'BKG-']
                for pattern in internal_patterns:
                    if pattern.lower() in hcn_lower:
                        is_invalid = True
                        result['reason'] = f"Rejected: '{hcn_str}' looks like internal ref. " + result.get('reason', '')
                        break
                
                if is_invalid:
                    result['hcn_number'] = None
                    if result.get('category') == 'Received':
                        result['category'] = 'Non Critical'
            
            # If Critical, HCN must be null
            if result.get('category') == 'Critical' and result.get('hcn_number'):
                result['hcn_number'] = None
            
            return {
                'hcn': result.get('hcn_number'),
                'category': result.get('category', 'Non Critical'),
                'reason': result.get('reason', '')
            }
            
        except Exception as e:
            print(f"      ⚠️ OpenAI error: {str(e)}")
            return {'hcn': None, 'category': 'Non Critical', 'reason': 'Analysis failed'}
    
    def find_matching_booking(self, df, subject, body):
        """Find which booking the email is about"""
        text = f"{subject} {body}".lower()
        
        for idx, row in df.iterrows():
            file_no = str(row.get('FileNo', ''))
            supplier_ref = str(row.get('SupplierRef', ''))
            guest_name = str(row.get('GuestName', ''))
            
            if file_no and file_no.lower() in text:
                return idx
            if supplier_ref and pd.notna(row.get('SupplierRef')) and supplier_ref.lower() in text:
                return idx
            if guest_name and len(guest_name) > 5:
                clean_name = re.sub(r'^(MR\.|MRS\.|MS\.)\s*', '', guest_name, flags=re.IGNORECASE)
                if clean_name.lower() in text:
                    return idx
        return None
    
    # ==================== MAIN PROCESS ====================
    
    def process_all(self):
        """
        Main process:
        1. Send initial emails to new bookings
        2. Check inbox and analyze replies with OpenAI
        3. Auto-send reminders if no HCN after 2 hours
        """
        print("\n" + "="*60)
        print("HCN EMAIL MANAGEMENT - PROCESSING")
        print("="*60)
        
        df = self.read_excel()
        now = datetime.now()
        reminder_threshold = now - timedelta(hours=REMINDER_AFTER_HOURS)
        
        # Filter relevant bookings (Confirmed/Vouchered)
        df['Status_lower'] = df['Status'].str.lower().str.strip()
        
        # ========== STEP 1: SEND INITIAL EMAILS ==========
        print("\n" + "-"*60)
        print("📤 STEP 1: Checking for new bookings to email...")
        print("-"*60)
        
        # Find bookings that need initial email
        new_bookings = df[
            (df['Status_lower'].isin(['confirmed', 'vouchered'])) &
            (pd.isna(df['SupplierHCN']) | (df['SupplierHCN'].astype(str).str.strip() == '')) &
            (df['EmailSent'] != 'Yes')
        ]
        
        initial_sent = 0
        if len(new_bookings) > 0:
            print(f"   Found {len(new_bookings)} new bookings to email")
            
            for idx in new_bookings.index:
                row = df.loc[idx]
                recipient = self.get_recipient_email(row)
                
                if not recipient:
                    continue
                
                subject, body = self.create_email_content(row, is_reminder=False)
                success, msg = self.send_email(recipient, subject, body)
                
                if success:
                    df.at[idx, 'EmailSent'] = 'Yes'
                    df.at[idx, 'EmailSentTime'] = now.strftime('%Y-%m-%d %H:%M:%S')
                    initial_sent += 1
                    print(f"   [✓] {row.get('FileNo')} | {row.get('GuestName')} -> {recipient}")
                else:
                    print(f"   [✗] {row.get('FileNo')}: {msg}")
                
                time.sleep(DELAY_BETWEEN_EMAILS)
            
            print(f"   ✅ Sent {initial_sent} initial emails")
        else:
            print("   No new bookings to email")
        
        # ========== STEP 2: CHECK INBOX & ANALYZE ==========
        print("\n" + "-"*60)
        print("📥 STEP 2: Checking inbox for replies...")
        print("-"*60)
        
        mail = self.connect_gmail_imap()
        replies_processed = {'Received': 0, 'Critical': 0, 'Non Critical': 0}
        
        if mail:
            date_since = (now - timedelta(days=DAYS_TO_CHECK)).strftime('%d-%b-%Y')
            _, nums = mail.search(None, f'(SINCE {date_since})')
            email_ids = nums[0].split()
            print(f"   Checking {len(email_ids)} emails...")
            
            processed = set()
            
            for num in email_ids:
                try:
                    _, data = mail.fetch(num, '(RFC822)')
                    msg = email.message_from_bytes(data[0][1])
                    
                    subject = self.decode_header_value(msg['Subject'])
                    body = self.get_email_body(msg)
                    
                    if not body.strip():
                        continue
                    
                    match_idx = self.find_matching_booking(df, subject, body)
                    
                    if match_idx is not None and match_idx not in processed:
                        row = df.loc[match_idx]
                        
                        # Skip if already has Issue (already processed)
                        if pd.notna(row.get('Issue')) and str(row.get('Issue')).strip():
                            continue
                        
                        print(f"\n   📩 Reply found: {row.get('FileNo')} | {row.get('GuestName')}")
                        print(f"      🤖 Analyzing with OpenAI...")
                        
                        booking_info = {
                            'guest_name': row.get('GuestName'),
                            'hotel_name': row.get('HotelName'),
                            'file_no': row.get('FileNo'),
                            'supplier_ref': row.get('SupplierRef', '')
                        }
                        
                        analysis = self.analyze_with_openai(subject, body, booking_info)
                        category = analysis['category']
                        
                        if category == 'Received' and analysis['hcn']:
                            df.at[match_idx, 'SupplierHCN'] = analysis['hcn']
                            df.at[match_idx, 'Issue'] = 'Received'
                            print(f"      ✅ RECEIVED - HCN: {analysis['hcn']}")
                            replies_processed['Received'] += 1
                        elif category == 'Critical':
                            df.at[match_idx, 'Issue'] = 'Critical'
                            print(f"      🚨 CRITICAL - {analysis['reason']}")
                            replies_processed['Critical'] += 1
                        else:
                            df.at[match_idx, 'Issue'] = 'Non Critical'
                            print(f"      ℹ️  NON CRITICAL - {analysis['reason']}")
                            replies_processed['Non Critical'] += 1
                        
                        processed.add(match_idx)
                        
                except Exception as e:
                    continue
            
            mail.logout()
            
            total_replies = sum(replies_processed.values())
            print(f"\n   ✅ Processed {total_replies} replies")
        else:
            print("   ❌ Could not connect to Gmail")
        
        # ========== STEP 3: SEND REMINDERS ==========
        print("\n" + "-"*60)
        print("🔔 STEP 3: Checking for reminders (2+ hours, no HCN)...")
        print("-"*60)
        
        reminders_sent = 0
        
        for idx, row in df.iterrows():
            # Conditions for reminder:
            # 1. Confirmed/Vouchered status
            # 2. Email was sent
            # 3. No HCN received (Issue is empty/Pending OR Non Critical)
            # 4. NOT Critical (needs manual action)
            # 5. NOT already Received
            # 6. Reminder not sent yet
            # 7. 2+ hours passed since initial email
            
            status = str(row.get('Status', '')).lower().strip()
            if status not in ['confirmed', 'vouchered']:
                continue
            
            if row.get('EmailSent') != 'Yes':
                continue
            
            # Skip if already got HCN or has Critical issue
            issue = row.get('Issue')
            if pd.notna(issue):
                issue_str = str(issue).strip()
                if issue_str == 'Received':
                    continue  # Got HCN, no reminder needed
                if issue_str == 'Critical':
                    continue  # Critical issue, needs manual action
            
            # Skip if already reminded
            if row.get('ReminderSent') == 'Yes':
                continue
            
            # Check if 2+ hours passed
            sent_time_str = row.get('EmailSentTime')
            if pd.isna(sent_time_str) or str(sent_time_str).strip() == '':
                continue
            
            try:
                sent_time = datetime.strptime(str(sent_time_str), '%Y-%m-%d %H:%M:%S')
                if sent_time > reminder_threshold:
                    continue  # Less than 2 hours, skip
            except:
                continue
            
            # Send reminder
            recipient = self.get_recipient_email(row)
            if not recipient:
                continue
            
            subject, body = self.create_email_content(row, is_reminder=True)
            success, msg = self.send_email(recipient, subject, body)
            
            if success:
                df.at[idx, 'ReminderSent'] = 'Yes'
                df.at[idx, 'ReminderTime'] = now.strftime('%Y-%m-%d %H:%M:%S')
                reminders_sent += 1
                issue_status = str(row.get('Issue')).strip() if pd.notna(row.get('Issue')) else "Pending"
                print(f"   [✓] REMINDER: {row.get('FileNo')} | {row.get('GuestName')} | Was: {issue_status}")
            
            time.sleep(DELAY_BETWEEN_EMAILS)
        
        if reminders_sent > 0:
            print(f"\n   ✅ Sent {reminders_sent} reminders")
        else:
            print("   No reminders needed (all within 2 hours or already reminded)")
        
        # ========== SAVE & SUMMARY ==========
        self.save_excel(df)
        
        # Show summary
        print("\n" + "="*60)
        print("📊 SUMMARY")
        print("="*60)
        
        relevant = df[df['Status_lower'].isin(['confirmed', 'vouchered'])]
        
        received = (relevant['Issue'] == 'Received').sum()
        critical = (relevant['Issue'] == 'Critical').sum()
        non_critical = (relevant['Issue'] == 'Non Critical').sum()
        pending = relevant['Issue'].isna().sum()
        reminded = (relevant['ReminderSent'] == 'Yes').sum()
        
        print(f"\nThis Run:")
        print(f"   📤 Initial emails sent: {initial_sent}")
        print(f"   📥 Replies processed: {sum(replies_processed.values())}")
        print(f"   🔔 Reminders sent: {reminders_sent}")
        
        print(f"\nOverall Status:")
        print(f"   ✅ Received (HCN): {received}")
        print(f"   🚨 Critical: {critical}")
        print(f"   ℹ️  Non Critical: {non_critical}")
        print(f"   ⏳ Pending: {pending}")
        print(f"   🔔 Total reminded: {reminded}")
        
        # Show critical issues
        critical_rows = relevant[relevant['Issue'] == 'Critical']
        if len(critical_rows) > 0:
            print(f"\n🚨 CRITICAL ISSUES - Need Attention:")
            for _, row in critical_rows.iterrows():
                print(f"   • {row.get('FileNo')} | {row.get('GuestName')} | {row.get('HotelName')}")
        
        print("\n" + "="*60)
        print("✅ PROCESS COMPLETE")
        print("="*60)
    
    def get_summary_stats(self):
        """Get summary statistics as a dictionary (for API)"""
        df = self.read_excel()
        df['Status_lower'] = df['Status'].str.lower().str.strip()
        relevant = df[df['Status_lower'].isin(['confirmed', 'vouchered'])]

        total = len(relevant)
        emailed = int((relevant['EmailSent'] == 'Yes').sum())
        reminded = int((relevant['ReminderSent'] == 'Yes').sum())

        received = int((relevant['Issue'] == 'Received').sum())
        critical = int((relevant['Issue'] == 'Critical').sum())
        non_critical = int((relevant['Issue'] == 'Non Critical').sum())
        pending = int(relevant['Issue'].isna().sum())

        return {
            'total': total,
            'received': received,
            'critical': critical,
            'non_critical': non_critical,
            'pending': pending,
            'emailed': emailed,
            'reminded': reminded
        }

    def show_status(self):
        """Show current status only"""
        print("\n" + "="*60)
        print("📊 STATUS OVERVIEW")
        print("="*60)

        df = self.read_excel()
        df['Status_lower'] = df['Status'].str.lower().str.strip()
        relevant = df[df['Status_lower'].isin(['confirmed', 'vouchered'])]

        total = len(relevant)
        emailed = (relevant['EmailSent'] == 'Yes').sum()
        reminded = (relevant['ReminderSent'] == 'Yes').sum()

        received = (relevant['Issue'] == 'Received').sum()
        critical = (relevant['Issue'] == 'Critical').sum()
        non_critical = (relevant['Issue'] == 'Non Critical').sum()
        pending = relevant['Issue'].isna().sum()

        print(f"\nTotal Confirmed/Vouchered: {total}")
        print(f"\n📤 Emails:")
        print(f"   Initial sent: {emailed}")
        print(f"   Reminders sent: {reminded}")

        print(f"\n📥 Status (via OpenAI):")
        print(f"   ✅ Received (HCN): {received}")
        print(f"   🚨 Critical: {critical}")
        print(f"   ℹ️  Non Critical: {non_critical}")
        print(f"   ⏳ Pending: {pending}")

        # Show details
        if critical > 0:
            print(f"\n🚨 CRITICAL ISSUES:")
            for _, row in relevant[relevant['Issue'] == 'Critical'].iterrows():
                print(f"   • {row.get('FileNo')} | {row.get('GuestName')} | {row.get('HotelName')}")

        if pending > 0:
            print(f"\n⏳ PENDING (Awaiting reply):")
            pending_rows = relevant[relevant['Issue'].isna()]
            for _, row in pending_rows.head(10).iterrows():
                reminded_status = "✓ Reminded" if row.get('ReminderSent') == 'Yes' else "Not reminded yet"
                print(f"   • {row.get('FileNo')} | {row.get('GuestName')} | {reminded_status}")
            if len(pending_rows) > 10:
                print(f"   ... and {len(pending_rows) - 10} more")


def main():
    print("="*60)
    print("HCN EMAIL MANAGEMENT SYSTEM")
    print("="*60)

    # Validate configuration
    from config import validate_config
    errors = validate_config()

    if errors:
        print("\n⚠️  Configuration Errors:")
        for error in errors:
            print(f"   - {error}")
        print("\n💡 Please update the .env file with correct values")
        print("   See .env.example for reference")
        return
    
    manager = HCNEmailManager()
    
    while True:
        print("\n" + "-"*40)
        print("MENU:")
        print("1. Run Process (Send → Check → Remind)")
        print("2. Show Status")
        print("3. Exit")
        print("-"*40)
        
        choice = input("Choice (1-3): ").strip()
        
        if choice == '1':
            confirm = input("Run full process? (yes/no): ").strip().lower()
            if confirm == 'yes':
                manager.process_all()
        elif choice == '2':
            manager.show_status()
        elif choice == '3':
            print("Goodbye!")
            break
        else:
            print("Invalid choice")


if __name__ == "__main__":
    main()